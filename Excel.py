import os
from openpyxl import Workbook, load_workbook

class ExcelManager:
    def __init__(self, file_path: str, column_titles: list):
        self.file_path = file_path
        self.column_titles = column_titles
        self.sheet_titles = []

        # Vérification de l'existence du fichier
        if os.path.exists(file_path):
            self.workbook = load_workbook(file_path)
            self.sheet_titles = self.workbook.sheetnames
        else:
            # Création du fichier et de la première feuille par défaut
            self.workbook = Workbook()
            self.workbook.active.title = "Sheet1"  # Renommer la feuille par défaut
            self.save()

    def save(self):
        """Enregistre le fichier Excel."""
        self.workbook.save(self.file_path)

    def add_row(self, sheet_name: str, row_data: list):
        """Ajoute une nouvelle ligne à la feuille spécifiée. Si la feuille n'existe pas, elle est créée avec les titres de colonne."""
        if sheet_name not in self.sheet_titles:
            sheet = self.workbook.create_sheet(title=sheet_name)
            sheet.append(self.column_titles)
            self.sheet_titles.append(sheet_name)
        else:
            sheet = self.workbook[sheet_name]

        # Traitement des données de la ligne : remplacer '.' par ',' uniquement pour les floats avec un point
        processed_row = [
            str(item).replace('.', ',') if isinstance(item, tuple) else str(item).replace('.',',') if self.is_float_and_contains_dot(
                item) else item
            for item in row_data
        ]
        sheet.append(processed_row)
        self.save()

    def get_last_row_first_column(self, sheet_name: str) -> int:
        """Vérifie la dernière ligne du sheet spécifié et renvoie la première colonne de cette ligne en int.
        Si la conversion échoue ou si le sheet n'existe pas, retourne 0."""

        # Si la feuille n'existe pas, retourner 0
        if sheet_name not in self.sheet_titles:
            return 0

        # Récupérer la feuille
        sheet = self.workbook[sheet_name]

        # Si la feuille est vide ou ne contient que l'en-tête, retourner 0
        if sheet.max_row < 2:  # max_row donne la dernière ligne non vide, 1 pour en-tête
            return 0

        # Récupérer la valeur de la première colonne de la dernière ligne
        last_row_value = sheet.cell(row=sheet.max_row, column=1).value

        # Tenter de convertir la valeur en entier, retourner 0 en cas d'échec
        try:
            return int(last_row_value)
        except (ValueError, TypeError):
            return 0

    def count_rows(self, sheet_name: str) -> int:
        """Renvoie le nombre total de lignes dans la feuille spécifiée, y compris les en-têtes."""
        if sheet_name in self.sheet_titles:
            sheet = self.workbook[sheet_name]
            return sheet.max_row  # max_row retourne le nombre de lignes non vides
        return 0

    def read_rows(self, sheet_name: str):
        """Parcourt les lignes de la feuille spécifiée et renvoie les données sous forme de liste de listes.
        Si la feuille n'existe pas, une erreur est générée si le fichier n'existe pas."""
        # Vérification de l'existence du fichier
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Le fichier {self.file_path} n'existe pas.")

        # Vérification de l'existence de la feuille
        if sheet_name not in self.sheet_titles:
            for sheet in self.workbook.sheetnames:
                print(f"Feuille : {sheet}")
            raise ValueError(f"La feuille {sheet_name} n'existe pas dans le fichier {self.file_path}.")

        # Récupérer la feuille
        sheet = self.workbook[sheet_name]

        # Parcourir les lignes et les ajouter à une liste
        rows = []
        for row in sheet.iter_rows(values_only=True):  # values_only=True pour obtenir uniquement les valeurs
            rows.append(row)

        return rows


    @staticmethod
    def is_float_and_contains_dot(value):
        """Vérifie si une valeur est un float (ou peut être convertie en float) et contient un point décimal."""
        try:
            # On essaie d'analyser la valeur pour voir si elle est un float
            return isinstance(value, (float, int)) or (isinstance(value, str) and '.' in value and float(value))
        except ValueError:
            return False
